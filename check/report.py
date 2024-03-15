import sys

from openpyxl import load_workbook
from openpyxl import Workbook

# Load the promo.xlsx file
promo_workbook = load_workbook('promo.xlsx')
promo_sheet = promo_workbook.active

# Load the bot.xlsx file
bot_workbook = load_workbook('bot.xlsx')
bot_sheet = bot_workbook.active

output_workbook = Workbook()
output_sheet = output_workbook.active

# Add headers to the output sheet
output_sheet.append(["Column1", "Column2", "Column3", "Column4", "Column5"])

# Iterate over the rows in bot.xlsx
for i, bot_row in enumerate(bot_sheet.iter_rows(min_row=2, values_only=True), start=2):
    bot_value = str(bot_row[3])  # Assuming the value is in the third column
    exists = any(bot_value == str(promo_row[0]) for promo_row in promo_sheet.iter_rows(min_row=2, values_only=True))
    sys.stdout.write(str(i) + "\n")
    sys.stdout.flush()
    if exists:
        output_sheet.append(bot_row)
output_workbook.save("out.xlsx")
