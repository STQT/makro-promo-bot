import requests
from openpyxl import Workbook
from openpyxl.reader.excel import load_workbook
import json


# Function to load progress from JSON file
def load_progress():
    try:
        with open('progress.json', 'r') as file:
            return json.load(file)
    except FileNotFoundError:
        return {}


# Function to save progress to JSON file
def save_progress(progress):
    with open('progress.json', 'w') as file:
        json.dump(progress, file)


# Load the promo.xlsx file
promo_workbook = load_workbook('promo.xlsx')
promo_sheet = promo_workbook.active

# Create a new workbook and worksheet
output_workbook = Workbook()
output_sheet = output_workbook.active

# Add headers to the output sheet
output_sheet.append(["Phone", "Fullname", "Promotion", "Code", "Created At"])

progress = load_progress()
token = "23342e36c2c0a163cca9f4874372a36331073f16"

for num, promo_row in enumerate(promo_sheet.iter_rows(min_row=2, values_only=True)):
    promo_value = str(promo_row[0])
    if promo_value in progress:
        continue  # Skip if already processed

    api_url = f'https://bot.makromarket.uz/promo/{promo_value}/'
    headers = {'Authorization': f'Token {token}'}


    response = requests.get(api_url, headers=headers)
    print(response, promo_row, num)
    if response.status_code == 200:
        data = response.json()
        user = data["user"]
        if "phone" in user and "fullname" in user:
            output_sheet.append([
                user["phone"],
                user["fullname"],
                data["promotion"],
                data["code"],
                data["created_at"]
            ])
            # Save processed data to JSON
            progress[promo_value] = {
                "Phone": user["phone"],
                "Fullname": user["fullname"],
                "Promotion": data["promotion"],
                "Code": data["code"],
                "Created At": data["created_at"]
            }
            save_progress(progress)
    elif response.status_code == 404:
        progress[promo_value] = True
        save_progress(progress)
    else:
        raise Exception(f"Toxtadi {response.text}")


# Save the output workbook
output_workbook.save("output.xlsx")
